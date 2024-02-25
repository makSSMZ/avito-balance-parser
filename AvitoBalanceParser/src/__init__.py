import requests
import pandas as pd
import openpyxl
from datetime import datetime, timedelta
import configparser
import shutil
import re

config = configparser.ConfigParser()
config.read(r'config\config.ini')

avito_link = config['SETTINGS']['AvitoLink']
excel_path = config['SETTINGS']['ExcelPath']
advance_cell_letter = config['SETTINGS']['AdvanceCellLetter']
applies_cell_letter = config['SETTINGS']['AppliesCellLetter']
ads_count_cell_letter = config['SETTINGS']['AdsCountCellLetter']
page_count = int(config['SETTINGS']['PageCount']) + 1


##########################
#DELETE Get balance of current user
def get_balance(token):
    headers = {"Authorization": "Bearer "+ token}
    resp = requests.get(avito_link + '/core/v1/accounts/{id}/balance/', headers=headers)
    print(resp.url)

#DELETE Get id of current user
def get_user_infoID(token):
    headers = {"Authorization": "Bearer "+ token}
    resp = requests.get(avito_link + '/core/v1/accounts/self', headers=headers)
    return resp.json()['id']

#DELETE
def get_stat(token):
    headers = {"Authorization": "Bearer "+ token}
    param = {"dateFrom" : "2023-10-15", "dateTo" : "2023-10-15", "fields" : ["views"], "itemIds" : [{id}] , "periodGrouping": "day"}
    resp = requests.post(avito_link + '/stats/v1/accounts/{id}/items', json=param, headers=headers)
    resp.json()['result']['items'][0]['stats'][0]['uniqViews']
    print(resp.json()['result']['items'][0]['itemId'])

##########################

#Request to get all ads from page
def get_ads(token, page):
    headers = {"Authorization": "Bearer "+ token}
    params = {"per_page": 100, "page": page}
    resp = requests.get(avito_link + '/core/v1/items', params=params, headers=headers)

    return len(resp.json()['resources'])

#Get all ads from all pages 
def count_all_ads(token):
    adsCount = 0
    i = 1
    while i < page_count:
        count = get_ads(token, i)
        adsCount+= count
        i+=1
        if count < 100:
            break

    return adsCount

#Get token for bearer
def get_token(client_id, client_secret):
    param = {'client_id': client_id, 'client_secret': client_secret, 'grant_type': 'client_credentials'}
    resp = requests.post(avito_link + '/token', data=param)
    print(resp.json()['access_token'])
    return resp.json()['access_token']

#Get advance of current user 
def get_advance(token):
    headers = {"Authorization": "Bearer " + token, "X-Source" : "Ajkaz"}
    param = {}
    resp = requests.post(avito_link + '/cpa/v2/balanceInfo', json=param, headers=headers)
    return  resp.json()['result']['balance'] / 100

#Get applies count of current user for prev day
def get_applies_count(token):
    headers = {"Authorization": "Bearer "+ token}
    date = (datetime.now()-timedelta(days=2)).strftime("%Y-%m-%d")
    params = {"updatedAtFrom": date}
    resp = requests.get(avito_link + '/job/v1/applications/get_ids', params=params, headers=headers)

    true_dates = [] 
    for date in resp.json()['applies']:
        true_date = datetime.strptime(re.sub("(?=\.).*", "", date['created_at']), '%Y-%m-%dT%H:%M:%S') + timedelta(hours=3)
        if (true_date.day == (datetime.now()-timedelta(days=1)).day):
            true_dates.append(true_date) 

    return len(true_dates)

#Search and write advance value in the excel
def write_advance_cell(acc_name, advance_value):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    for row in ws.iter_rows():
        for cell in row:
            if cell.value == acc_name:
                cell_numb = str(cell.row)
    
    ws[advance_cell_letter + cell_numb] = advance_value
    wb.save(excel_path)

#Search and write applies value in the excel
def write_applies_cell(acc_name, applies_value):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    for row in ws.iter_rows():
        for cell in row:
            if cell.value == acc_name:
                cell_numb = str(cell.row)
    
    ws[applies_cell_letter + cell_numb] = applies_value
    wb.save(excel_path)

#Search and write ads count value in the excel
def write_ads_cell(acc_name, ads_count_value):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    for row in ws.iter_rows():
        for cell in row:
            if cell.value == acc_name:
                cell_numb = str(cell.row)
    
    ws[ads_count_cell_letter + cell_numb] = ads_count_value
    wb.save(excel_path)   

#Maim func
def execute_main():
    df = pd.read_excel(excel_path, usecols='A, E, F')

    for index, row in df.iterrows():
        if pd.notna(row["clientId"]):
            try:
                print("Работаю с:")
                print(row["Номер Авито"])
                print(row["clientId"])
                print(row["clientSecret"])
        
                token = get_token(row["clientId"], row["clientSecret"])
                advance = get_advance(token)
                applies = get_applies_count(token)
                ads = count_all_ads(token)
                write_advance_cell(row["Номер Авито"], advance)
                write_applies_cell(row["Номер Авито"], applies)
                write_ads_cell(row["Номер Авито"], ads)
            except Exception as ex:
                try:
                    print("Возникла ошибка: " + ex.args[0])

                    write_advance_cell(row["Номер Авито"], "Ошибка")
                    write_applies_cell(row["Номер Авито"], "Ошибка")
                    write_ads_cell(row["Номер Авито"], "Ошибка")
                except Exception:
                    print("Не удалось записать в Excel!")


if __name__ == "__main__":
    try:
        shutil.copyfile(excel_path, "Copy_" + excel_path)
        print("Сделал копию файла")
        execute_main()
        input("Выполнение успешно завершено. Нажмите enter: ")
    except Exception as ex:
        print("Ошибка при выполнении: " + ex.args[0])
        input("Нажмите любую кноку: ")